import discord
from discord.ext import commands, tasks
import asyncio
import os
import csv
import io
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import sqlite3

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== CONFIG ====================
DISCORD_TOKEN         = os.environ.get("DISCORD_TOKEN")
APPROVAL_CHANNELS     = ["ssrp-approval", "outrider-ssrp-approval"]
REPORT_CHANNEL_NAME   = "leaderboard-ssrp-command"
INACTIVE_CHANNEL_NAME = "inactive-permission"
LOG_CHANNEL_NAME      = "ssrp-check"
ADMIN_ROLE_ID         = 1358301859663839374
SUBMIT_COOLDOWN       = 30  # detik jeda antar submit

# ─── EDIT TEKS PENGUMUMAN SENIN ──────────────────────
ANNOUNCE_TEXT = """📢 **PENGUMUMAN SSRP MINGGUAN!**

⚠️ Poin minggu ini telah **direset**!
Semangat submit SSRP minggu ini ya!!

🎭 Submit screenshot SSRP di channel ini!
🏆 Cek poin: `!point`
📊 Leaderboard: `!lb`"""
# ═════════════════════════════════════════════════════

# Auto export times WIB → UTC
# 00:00 WIB = 17:00 UTC (hari sebelumnya)
# 15:15 WIB = 08:15 UTC
# 20:00 WIB = 13:00 UTC
AUTO_EXPORT_TIMES_UTC = [(17, 0), (8, 15), (13, 0)]
# ================================================

intents = discord.Intents.default()
intents.message_content = True
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)
last_submit: dict[str, datetime] = defaultdict(lambda: datetime.min.replace(tzinfo=timezone.utc))
pending_removals: dict[str, dict] = {}

# ─── Database ─────────────────────────────────────────────────────────────────
DB = "/data/ssrp_points.db"

def init_db():
    os.makedirs("/data", exist_ok=True)
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS points (
        user_id     TEXT PRIMARY KEY,
        username    TEXT,
        total_points INTEGER DEFAULT 0,
        week_points  INTEGER DEFAULT 0,
        last_reset   TEXT,
        last_submit  TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS submissions (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id      TEXT,
        username     TEXT,
        submitted_at TEXT,
        photo_count  INTEGER,
        points_given INTEGER DEFAULT 0
    )""")
    conn.commit()
    conn.close()

def upsert_user(user_id, username):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        c.execute("""INSERT INTO points
            (user_id, username, total_points, week_points, last_reset, last_submit)
            VALUES (?, ?, 0, 0, ?, NULL)""",
            (user_id, username, datetime.now(timezone.utc).isoformat()))
    conn.commit()
    conn.close()

def add_points(user_id, username, photo_count):
    """Tambah poin sebanyak jumlah foto."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    now = datetime.now(timezone.utc)

    c.execute("SELECT week_points, last_reset FROM points WHERE user_id=?", (user_id,))
    row = c.fetchone()
    if row:
        last_reset = datetime.fromisoformat(row[1])
        if last_reset.tzinfo is None:
            last_reset = last_reset.replace(tzinfo=timezone.utc)
        if now - last_reset > timedelta(days=7):
            c.execute("UPDATE points SET week_points=0, last_reset=? WHERE user_id=?",
                      (now.isoformat(), user_id))

    c.execute("""UPDATE points SET
        total_points = total_points + ?,
        week_points  = week_points + ?,
        username     = ?,
        last_submit  = ?
        WHERE user_id = ?""",
        (photo_count, photo_count, username, now.isoformat(), user_id))
    conn.commit()

    c.execute("SELECT total_points, week_points FROM points WHERE user_id=?", (user_id,))
    result = c.fetchone()
    conn.close()
    return result

def log_submission(user_id, username, photo_count, points_given):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""INSERT INTO submissions
        (user_id, username, submitted_at, photo_count, points_given)
        VALUES (?, ?, ?, ?, ?)""",
        (user_id, username, datetime.now(timezone.utc).isoformat(), photo_count, points_given))
    conn.commit()
    conn.close()

def get_all_points():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""SELECT user_id, username, total_points, week_points, last_submit
                 FROM points ORDER BY total_points DESC""")
    rows = c.fetchall()
    conn.close()
    return rows

def get_user_point(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT total_points, week_points, last_submit FROM points WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row

def get_user_rank(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points ORDER BY total_points DESC")
    rows = c.fetchall()
    conn.close()
    for i, (uid,) in enumerate(rows, 1):
        if uid == user_id:
            return i, len(rows)
    return None, None

def get_inactive_users(days=7):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    threshold = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    c.execute("""SELECT user_id, username, last_submit FROM points
                 WHERE last_submit IS NULL OR last_submit < ?""", (threshold,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_user_photo_count(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT COALESCE(SUM(photo_count), 0) FROM submissions WHERE user_id=?", (user_id,))
    result = c.fetchone()[0]
    conn.close()
    return result

def get_user_submit_count(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM submissions WHERE user_id=?", (user_id,))
    result = c.fetchone()[0]
    conn.close()
    return result

def manual_add_point(user_id, username, amount):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        c.execute("""INSERT INTO points
            (user_id, username, total_points, week_points, last_reset, last_submit)
            VALUES (?, ?, 0, 0, ?, NULL)""",
            (user_id, username, datetime.now(timezone.utc).isoformat()))
    c.execute("""UPDATE points SET
        total_points = total_points + ?,
        week_points  = week_points + ?,
        username     = ?
        WHERE user_id = ?""", (amount, amount, username, user_id))
    conn.commit()
    c.execute("SELECT total_points, week_points FROM points WHERE user_id=?", (user_id,))
    result = c.fetchone()
    conn.close()
    return result

def manual_remove_point(user_id, amount):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT total_points, week_points FROM points WHERE user_id=?", (user_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return None
    new_total = max(0, row[0] - amount)
    new_week  = max(0, row[1] - amount)
    c.execute("UPDATE points SET total_points=?, week_points=? WHERE user_id=?",
              (new_total, new_week, user_id))
    conn.commit()
    conn.close()
    return (new_total, new_week)

def db_reset_user(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        conn.close()
        return False
    c.execute("""UPDATE points SET
        total_points=0, week_points=0,
        last_reset=?, last_submit=NULL
        WHERE user_id=?""",
        (datetime.now(timezone.utc).isoformat(), user_id))
    conn.commit()
    conn.close()
    return True

def db_remove_user(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        conn.close()
        return False
    c.execute("DELETE FROM points WHERE user_id=?", (user_id,))
    c.execute("DELETE FROM submissions WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()
    return True

def reset_all_week_points():
    """Reset semua poin mingguan."""
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("UPDATE points SET week_points=0, last_reset=?",
              (datetime.now(timezone.utc).isoformat(),))
    conn.commit()
    conn.close()

def format_ts(ts_str):
    if not ts_str:
        return "-"
    try:
        dt = datetime.fromisoformat(ts_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return (dt + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M WIB")
    except:
        return ts_str

# ─── Excel & CSV ──────────────────────────────────────────────────────────────
def generate_excel(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SSRP Points"

    HDR_BG   = "1C3A1C"
    ROW_GRN  = "2D5A2D"
    ROW_ALT  = "245024"
    ROW_RED  = "8B0000"
    WHITE    = "FFFFFF"

    thin = Border(
        left=Side(style="thin", color="3A6B3A"),
        right=Side(style="thin", color="3A6B3A"),
        top=Side(style="thin", color="3A6B3A"),
        bottom=Side(style="thin", color="3A6B3A")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    headers    = ["No", "Username", "User ID", "Total Poin", "Poin Minggu Ini", "Total Foto", "Terakhir Submit"]
    col_widths = [5,    28,         22,         14,           18,                12,            25]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.font      = Font(bold=True, color=WHITE, name="Arial", size=11)
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    for i, (uid, uname, total, week, last_sub) in enumerate(rows, 1):
        r       = i + 1
        is_zero = (total == 0)
        bg      = ROW_RED if is_zero else (ROW_ALT if i % 2 == 0 else ROW_GRN)
        photos  = get_user_photo_count(uid)
        vals    = [i, uname, uid, total, week, photos, format_ts(last_sub)]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", size=10, color=WHITE, bold=is_zero)
            cell.border    = thin
            cell.alignment = center if col in (1, 4, 5, 6) else left
        ws.row_dimensions[r].height = 22

    tr = len(rows) + 2
    ws.row_dimensions[tr].height = 26
    for col, val in enumerate(["TOTAL","","",
        f"=SUM(D2:D{len(rows)+1})",
        f"=SUM(E2:E{len(rows)+1})",
        f"=SUM(F2:F{len(rows)+1})", ""], 1):
        cell = ws.cell(row=tr, column=col, value=val)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.font      = Font(bold=True, color=WHITE, name="Arial", size=11)
        cell.alignment = center
        cell.border    = thin

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def generate_csv(rows: list) -> bytes:
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["No","Username","User ID","Total Poin","Poin Minggu Ini","Total Foto","Terakhir Submit"])
    for i, (uid, uname, total, week, last_sub) in enumerate(rows, 1):
        w.writerow([i, uname, uid, total, week, get_user_photo_count(uid), format_ts(last_sub)])
    return buf.getvalue().encode("utf-8-sig")

# ─── Leaderboard ──────────────────────────────────────────────────────────────
PER_PAGE = 15

def build_lb_embed(rows, page=0):
    start  = page * PER_PAGE
    chunk  = rows[start:start + PER_PAGE]
    total  = len(rows)
    now    = (datetime.now(timezone.utc) + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M")
    ICONS  = {1:"🥇", 2:"🥈", 3:"🥉"}

    embed = discord.Embed(title="🏆 SSRP Leaderboard", color=0xF1C40F,
                          timestamp=datetime.now(timezone.utc))
    embed.set_footer(text=f"Halaman {page+1} • Total {total} member • {now} WIB")

    if not chunk:
        embed.description = "_Belum ada data._"
        return embed

    desc = ""
    for i, (uid, uname, total_p, week, _) in enumerate(chunk, start + 1):
        icon = ICONS.get(i, f"`#{i}`")
        desc += f"{icon} **{uname}** — {week} poin minggu ini _(total: {total_p})_\n"
    embed.description = desc
    return embed

# ─── Leaderboard View — timeout=None ──────────────────────────────────────────
class LBView(discord.ui.View):
    """Timeout=None agar tombol tidak pernah expired."""
    def __init__(self, rows, page=0):
        super().__init__(timeout=None)
        self.rows     = rows
        self.page     = page
        self.max_page = max(0, (len(rows) - 1) // PER_PAGE)
        self._upd()

    def _upd(self):
        self.prev_btn.disabled = (self.page == 0)
        self.next_btn.disabled = (self.page >= self.max_page)

    @discord.ui.button(label="◀ Prev", style=discord.ButtonStyle.secondary,
                       custom_id="lb_prev")
    async def prev_btn(self, interaction: discord.Interaction, btn: discord.ui.Button):
        self.rows = get_all_points()
        self.max_page = max(0, (len(self.rows) - 1) // PER_PAGE)
        self.page = max(0, self.page - 1)
        self._upd()
        await interaction.response.edit_message(embed=build_lb_embed(self.rows, self.page), view=self)

    @discord.ui.button(label="🔄 Refresh", style=discord.ButtonStyle.primary,
                       custom_id="lb_refresh")
    async def refresh_btn(self, interaction: discord.Interaction, btn: discord.ui.Button):
        self.rows = get_all_points()
        self.max_page = max(0, (len(self.rows) - 1) // PER_PAGE)
        self._upd()
        await interaction.response.edit_message(embed=build_lb_embed(self.rows, self.page), view=self)

    @discord.ui.button(label="Next ▶", style=discord.ButtonStyle.secondary,
                       custom_id="lb_next")
    async def next_btn(self, interaction: discord.Interaction, btn: discord.ui.Button):
        self.rows = get_all_points()
        self.max_page = max(0, (len(self.rows) - 1) // PER_PAGE)
        self.page = min(self.max_page, self.page + 1)
        self._upd()
        await interaction.response.edit_message(embed=build_lb_embed(self.rows, self.page), view=self)

    @discord.ui.button(label="📊 Saya", style=discord.ButtonStyle.success,
                       custom_id="lb_me")
    async def me_btn(self, interaction: discord.Interaction, btn: discord.ui.Button):
        self.rows = get_all_points()
        uid = str(interaction.user.id)
        rank, total_m = get_user_rank(uid)
        if not rank:
            await interaction.response.send_message(
                "❌ Kamu belum pernah submit SSRP!", ephemeral=True)
            return
        self.page = (rank - 1) // PER_PAGE
        self.max_page = max(0, (len(self.rows) - 1) // PER_PAGE)
        self._upd()
        embed = build_lb_embed(self.rows, self.page)
        embed.set_author(
            name=f"📍 Kamu di posisi #{rank} dari {total_m} member",
            icon_url=interaction.user.display_avatar.url)
        await interaction.response.edit_message(embed=embed, view=self)

# ─── Tasks ────────────────────────────────────────────────────────────────────
@tasks.loop(minutes=1)
async def auto_reset_week():
    """Reset poin mingguan otomatis setiap Senin 00:00 WIB (Minggu 17:00 UTC)."""
    now = datetime.now(timezone.utc)
    if now.weekday() == 6 and now.hour == 17 and now.minute == 0:
        reset_all_week_points()
        for guild in bot.guilds:
            for ch_name in APPROVAL_CHANNELS:
                ch = discord.utils.get(guild.text_channels, name=ch_name)
                if ch:
                    try:
                        await ch.send(ANNOUNCE_TEXT)
                    except Exception:
                        pass
        print("✅ Auto reset mingguan selesai!")

@tasks.loop(minutes=1)
async def auto_export():
    now = datetime.now(timezone.utc)
    if (now.hour, now.minute) not in AUTO_EXPORT_TIMES_UTC:
        return
    rows = get_all_points()
    if not rows:
        return
    wib      = now + timedelta(hours=7)
    filename = f"ssrp_data_{wib.strftime('%d%m%Y_%H%M')}_WIB.xlsx"
    xls      = generate_excel(rows)
    for guild in bot.guilds:
        ch = discord.utils.get(guild.text_channels, name=LOG_CHANNEL_NAME)
        if ch:
            embed = discord.Embed(
                title="📊 Auto Export SSRP Data",
                description=f"Export otomatis **{wib.strftime('%d/%m/%Y %H:%M')} WIB**\nTotal: **{len(rows)}** member",
                color=0x2ECC71, timestamp=now)
            embed.set_footer(text="SSRP Auto Export")
            try:
                await ch.send(embed=embed,
                              file=discord.File(io.BytesIO(xls), filename=filename))
            except Exception as e:
                print(f"⚠️ Auto export error: {e}")

@tasks.loop(hours=24)
async def check_inactive():
    now = datetime.now(timezone.utc)
    # Senin jam 09:00 UTC = Senin 16:00 WIB
    if now.weekday() != 0 or now.hour != 9:
        return
    inactive = get_inactive_users(days=7)
    for guild in bot.guilds:
        ch = discord.utils.get(guild.text_channels, name=INACTIVE_CHANNEL_NAME)
        if not ch:
            continue
        if not inactive:
            await ch.send("✅ **Semua member sudah submit SSRP minggu ini!**")
            continue
        chunks = [inactive[i:i+20] for i in range(0, len(inactive), 20)]
        for idx, chunk in enumerate(chunks):
            desc = ""
            for uid, uname, last_sub in chunk:
                m   = guild.get_member(int(uid))
                mention = m.mention if m else f"**{uname}**"
                desc += f"• {mention} — terakhir: {format_ts(last_sub)}\n"
            embed = discord.Embed(
                title=f"⚠️ Tidak Aktif SSRP 7 Hari — Part {idx+1}",
                description=desc, color=0xE74C3C, timestamp=now)
            embed.set_footer(text=f"Total: {len(inactive)} member")
            await ch.send(embed=embed)
        for uid, uname, _ in inactive:
            m = guild.get_member(int(uid))
            if m:
                try:
                    await m.send(
                        f"⚠️ **Hei {m.display_name}!**\n"
                        f"Kamu belum submit SSRP dalam **7 hari terakhir**.\n"
                        f"Jangan lupa submit ya! 🎭")
                except Exception:
                    pass

# ─── Helper log ───────────────────────────────────────────────────────────────
async def send_admin_log(guild, action, admin, target, amount, before, after):
    ch = discord.utils.get(guild.text_channels, name=LOG_CHANNEL_NAME)
    if not ch:
        return
    color = 0x2ECC71 if action == "ADD" else 0xE74C3C
    icon  = "➕" if action == "ADD" else "➖"
    embed = discord.Embed(title=f"{icon} Poin {action} — Log Admin",
                          color=color, timestamp=datetime.now(timezone.utc))
    embed.add_field(name="👮 Admin",   value=admin.mention,    inline=True)
    embed.add_field(name="👤 Target",  value=target.mention,   inline=True)
    embed.add_field(name="🔢 Jumlah",  value=f"{amount} poin", inline=True)
    embed.add_field(name="📉 Sebelum", value=f"{before} poin", inline=True)
    embed.add_field(name="📈 Sesudah", value=f"{after} poin",  inline=True)
    embed.set_footer(text="SSRP Admin Log")
    await ch.send(embed=embed)

# ─── Events ───────────────────────────────────────────────────────────────────
@bot.event
async def on_ready():
    init_db()
    auto_reset_week.start()
    auto_export.start()
    check_inactive.start()
    print(f"✅ Bot online: {bot.user}")
    print(f"📋 Monitoring: {', '.join(['#'+c for c in APPROVAL_CHANNELS])}")

@bot.event
async def on_message(message: discord.Message):
    if message.author.bot:
        return
    if message.channel.name not in APPROVAL_CHANNELS:
        await bot.process_commands(message)
        return

    images = [a for a in message.attachments
              if a.content_type and a.content_type.startswith("image/")]
    if not images:
        await bot.process_commands(message)
        return

    user    = message.author
    user_id = str(user.id)
    now     = datetime.now(timezone.utc)
    elapsed = (now - last_submit[user_id]).total_seconds()

    if elapsed < SUBMIT_COOLDOWN:
        remaining = int(SUBMIT_COOLDOWN - elapsed)
        await message.add_reaction("⏳")
        await message.reply(
            f"⚠️ **{user.display_name}**, tunggu **{remaining} detik** lagi!\n"
            f"_(Submit ini tidak dihitung)_",
            delete_after=10)
        return

    last_submit[user_id] = now
    upsert_user(user_id, user.display_name)

    # 1 foto = 1 poin
    photo_count         = len(images)
    total_pts, week_pts = add_points(user_id, user.display_name, photo_count)
    log_submission(user_id, user.display_name, photo_count, photo_count)

    await message.add_reaction("✅")

    report_ch     = discord.utils.get(message.guild.text_channels, name=REPORT_CHANNEL_NAME) or message.channel
    msg_link      = f"https://discord.com/channels/{message.guild.id}/{message.channel.id}/{message.id}"
    channel_label = "Outrider SSRP" if message.channel.name == "outrider-ssrp-approval" else "SSRP"
    rank, total_m = get_user_rank(user_id)

    embed = discord.Embed(title=f"🎭 Laporan {channel_label} Baru!", color=0x57F287, timestamp=now)
    embed.set_author(name=user.display_name, icon_url=user.display_avatar.url)
    embed.add_field(name="👤 Player",           value=user.mention,                    inline=True)
    embed.add_field(name="🖼️ Jumlah Foto",      value=f"{photo_count} foto",           inline=True)
    embed.add_field(name="📅 Waktu",            value=f"<t:{int(now.timestamp())}:F>", inline=True)
    embed.add_field(name="➕ Poin Didapat",     value=f"**+{photo_count} poin**",      inline=True)
    embed.add_field(name="📈 Poin Minggu Ini",  value=f"**{week_pts} poin**",          inline=True)
    embed.add_field(name="🏆 Total Poin",       value=f"**{total_pts} poin**",         inline=True)
    embed.add_field(name="🏅 Ranking",          value=f"**#{rank}** dari {total_m}",   inline=True)
    embed.add_field(name="🔗 Link Bukti",       value=f"[Klik di sini]({msg_link})",   inline=True)
    embed.set_image(url=images[0].url)
    embed.set_footer(text=f"SSRP Checker • #{message.channel.name} • 1 foto = 1 poin")
    await report_ch.send(embed=embed)
    await bot.process_commands(message)

# ─── Commands ─────────────────────────────────────────────────────────────────
@bot.command(name="point")
async def cmd_point(ctx, member: discord.Member = None):
    target = member or ctx.author
    row    = get_user_point(str(target.id))
    if not row:
        await ctx.reply(f"**{target.display_name}** belum pernah submit SSRP.")
        return
    total, week, last_sub = row
    rank, total_m = get_user_rank(str(target.id))
    photos  = get_user_photo_count(str(target.id))
    submits = get_user_submit_count(str(target.id))
    last_str = f"<t:{int(datetime.fromisoformat(last_sub).timestamp())}:R>" if last_sub else "-"

    embed = discord.Embed(title="📊 SSRP Points", color=0x5865F2)
    embed.set_author(name=target.display_name, icon_url=target.display_avatar.url)
    embed.set_thumbnail(url=target.display_avatar.url)
    embed.add_field(name="🏅 Ranking",        value=f"**#{rank}** dari {total_m}", inline=True)
    embed.add_field(name="🏆 Total Poin",      value=f"**{total}**",                inline=True)
    embed.add_field(name="📈 Poin Minggu Ini", value=f"**{week}**",                 inline=True)
    embed.add_field(name="🖼️ Total Foto",      value=f"**{photos}**",               inline=True)
    embed.add_field(name="📋 Total Submit",    value=f"**{submits}x**",             inline=True)
    embed.add_field(name="🕐 Terakhir Submit", value=last_str,                      inline=True)
    await ctx.reply(embed=embed)

@bot.command(name="profile")
async def cmd_profile(ctx, member: discord.Member = None):
    target = member or ctx.author
    row    = get_user_point(str(target.id))
    if not row:
        await ctx.reply(f"**{target.display_name}** belum pernah submit SSRP.")
        return
    total, week, last_sub = row
    rank, total_m = get_user_rank(str(target.id))
    photos  = get_user_photo_count(str(target.id))
    submits = get_user_submit_count(str(target.id))
    status  = "🟢 Aktif minggu ini" if week > 0 else "🔴 Belum submit minggu ini"
    last_str = f"<t:{int(datetime.fromisoformat(last_sub).timestamp())}:F>" if last_sub else "-"

    embed = discord.Embed(title=f"🎭 Profil SSRP — {target.display_name}",
                          color=0x57F287, timestamp=datetime.now(timezone.utc))
    embed.set_thumbnail(url=target.display_avatar.url)
    embed.add_field(name="🏅 Ranking",        value=f"**#{rank}** dari {total_m}", inline=True)
    embed.add_field(name="🏆 Total Poin",      value=f"**{total}** poin",           inline=True)
    embed.add_field(name="📈 Poin Minggu Ini", value=f"**{week}** poin",            inline=True)
    embed.add_field(name="🖼️ Total Foto",      value=f"**{photos}** foto",          inline=True)
    embed.add_field(name="📋 Total Submit",    value=f"**{submits}x**",             inline=True)
    embed.add_field(name="🔵 Status",          value=status,                        inline=True)
    embed.add_field(name="🕐 Terakhir Submit", value=last_str,                      inline=False)
    embed.set_footer(text="SSRP Checker Bot")
    await ctx.reply(embed=embed)

@bot.command(name="leaderboard", aliases=["lb"])
async def cmd_lb(ctx):
    rows = get_all_points()
    if not rows:
        await ctx.reply("Belum ada data SSRP.")
        return
    view  = LBView(rows)
    embed = build_lb_embed(rows)
    await ctx.reply(embed=embed, view=view)

@bot.command(name="export")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_export(ctx, fmt: str = "excel"):
    rows = get_all_points()
    if not rows:
        await ctx.reply("Belum ada data.")
        return
    fmt   = fmt.lower()
    files = []
    if fmt in ("excel", "xlsx", "all"):
        files.append(discord.File(io.BytesIO(generate_excel(rows)), filename="ssrp_data.xlsx"))
    if fmt in ("csv", "all"):
        files.append(discord.File(io.BytesIO(generate_csv(rows)), filename="ssrp_data.csv"))
    if not files:
        await ctx.reply("Format tidak valid. Gunakan: `excel`, `csv`, atau `all`")
        return
    await ctx.reply(f"📊 Export SSRP — {len(rows)} member", files=files)

@bot.command(name="inactive")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_inactive(ctx, days: int = 7):
    inactive = get_inactive_users(days=days)
    if not inactive:
        await ctx.reply(f"✅ Semua member aktif dalam {days} hari terakhir!")
        return
    desc = ""
    for uid, uname, last_sub in inactive[:30]:
        m       = ctx.guild.get_member(int(uid))
        mention = m.mention if m else f"**{uname}**"
        desc   += f"• {mention} — {format_ts(last_sub)}\n"
    if len(inactive) > 30:
        desc += f"\n_...+{len(inactive)-30} lainnya. Gunakan `!export`._"
    embed = discord.Embed(title=f"⚠️ Tidak Aktif ({days} Hari)", description=desc, color=0xE74C3C)
    embed.set_footer(text=f"Total: {len(inactive)} member")
    await ctx.reply(embed=embed)

@bot.command(name="resetweek")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_resetweek(ctx):
    reset_all_week_points()
    await ctx.reply("✅ **Poin mingguan semua member sudah di-reset!**")

@bot.command(name="ekal")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_ekal(ctx):
    """Simulasi reset mingguan Senin 00:00 WIB — untuk testing."""
    reset_all_week_points()
    for ch_name in APPROVAL_CHANNELS:
        ch = discord.utils.get(ctx.guild.text_channels, name=ch_name)
        if ch:
            try:
                await ch.send(ANNOUNCE_TEXT)
            except Exception:
                pass
    await ctx.reply("🧪 **Simulasi reset mingguan berhasil!** Pengumuman sudah dikirim ke channel approval.")

@bot.command(name="resetuser")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_resetuser(ctx, member: discord.Member = None):
    if not member:
        await ctx.reply("❌ Gunakan: `!resetuser @user`")
        return
    if not db_reset_user(str(member.id)):
        await ctx.reply(f"❌ **{member.display_name}** tidak ada di database.")
        return
    embed = discord.Embed(title="🔄 User Di-reset!", color=0xE67E22)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player", value=member.mention,          inline=True)
    embed.add_field(name="🔄 Status", value="Semua poin jadi **0**", inline=True)
    embed.set_footer(text=f"Di-reset oleh {ctx.author.display_name}")
    await ctx.reply(embed=embed)
    ch = discord.utils.get(ctx.guild.text_channels, name=LOG_CHANNEL_NAME)
    if ch:
        log = discord.Embed(title="🔄 Reset User", color=0xE67E22, timestamp=datetime.now(timezone.utc))
        log.add_field(name="👮 Admin",  value=ctx.author.mention, inline=True)
        log.add_field(name="👤 Target", value=member.mention,     inline=True)
        await ch.send(embed=log)

@bot.command(name="removeuser")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_removeuser(ctx, member: discord.Member = None):
    if not member:
        await ctx.reply("❌ Gunakan: `!removeuser @user`")
        return
    if not db_remove_user(str(member.id)):
        await ctx.reply(f"❌ **{member.display_name}** tidak ada di database.")
        return
    embed = discord.Embed(title="🗑️ User Dihapus!", color=0xE74C3C)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player", value=member.mention,                        inline=True)
    embed.add_field(name="🗑️ Status", value="Dihapus dari database & leaderboard", inline=True)
    embed.set_footer(text=f"Dihapus oleh {ctx.author.display_name}")
    await ctx.reply(embed=embed)
    ch = discord.utils.get(ctx.guild.text_channels, name=LOG_CHANNEL_NAME)
    if ch:
        log = discord.Embed(title="🗑️ Remove User", color=0xE74C3C, timestamp=datetime.now(timezone.utc))
        log.add_field(name="👮 Admin",  value=ctx.author.mention, inline=True)
        log.add_field(name="👤 Target", value=member.mention,     inline=True)
        await ch.send(embed=log)

@bot.command(name="addpoint")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_addpoint(ctx, member: discord.Member = None, amount: int = 1):
    if not member:
        await ctx.reply("❌ Gunakan: `!addpoint @user [jumlah]`")
        return
    if amount < 1:
        await ctx.reply("❌ Jumlah minimal 1.")
        return
    before_row = get_user_point(str(member.id))
    before     = before_row[0] if before_row else 0
    result     = manual_add_point(str(member.id), member.display_name, amount)
    after      = result[0]
    embed = discord.Embed(title="➕ Poin Ditambahkan!", color=0x2ECC71)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player",     value=member.mention,   inline=True)
    embed.add_field(name="➕ Ditambah",   value=f"{amount} poin", inline=True)
    embed.add_field(name="🏆 Total Baru", value=f"{after} poin",  inline=True)
    await ctx.reply(embed=embed)
    await send_admin_log(ctx.guild, "ADD", ctx.author, member, amount, before, after)

@bot.command(name="removepoint")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_removepoint(ctx, member: discord.Member = None, amount: int = 1):
    if not member:
        await ctx.reply("❌ Gunakan: `!removepoint @user [jumlah]`")
        return
    if amount < 1:
        await ctx.reply("❌ Jumlah minimal 1.")
        return
    row = get_user_point(str(member.id))
    if not row:
        await ctx.reply(f"❌ **{member.display_name}** belum punya data poin.")
        return
    before = row[0]
    pending_removals[str(ctx.author.id)] = {
        "target_id": str(member.id),
        "target_obj": member,
        "amount": amount,
        "before": before,
        "expires": datetime.now(timezone.utc).timestamp() + 30
    }
    embed = discord.Embed(
        title="⚠️ Konfirmasi Hapus Poin",
        description=(
            f"Hapus **{amount} poin** dari {member.mention}?\n"
            f"Sekarang: **{before}** → setelah: **{max(0, before-amount)}**\n\n"
            f"Ketik `!confirm` dalam 30 detik atau `!cancel` untuk batal."
        ), color=0xE67E22)
    await ctx.reply(embed=embed)

@bot.command(name="confirm")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_confirm(ctx):
    admin_id = str(ctx.author.id)
    pending  = pending_removals.get(admin_id)
    if not pending:
        await ctx.reply("❌ Tidak ada pending removepoint.")
        return
    if datetime.now(timezone.utc).timestamp() > pending["expires"]:
        del pending_removals[admin_id]
        await ctx.reply("❌ Waktu habis. Ulangi `!removepoint`.")
        return
    result = manual_remove_point(pending["target_id"], pending["amount"])
    if not result:
        await ctx.reply("❌ Gagal, member tidak ditemukan.")
        return
    after  = result[0]
    member = pending["target_obj"]
    before = pending["before"]
    amount = pending["amount"]
    del pending_removals[admin_id]
    embed = discord.Embed(title="➖ Poin Dihapus!", color=0xE74C3C)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player",     value=member.mention,   inline=True)
    embed.add_field(name="➖ Dihapus",    value=f"{amount} poin", inline=True)
    embed.add_field(name="🏆 Total Baru", value=f"{after} poin",  inline=True)
    await ctx.reply(embed=embed)
    await send_admin_log(ctx.guild, "REMOVE", ctx.author, member, amount, before, after)

@bot.command(name="cancel")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cmd_cancel(ctx):
    if str(ctx.author.id) in pending_removals:
        del pending_removals[str(ctx.author.id)]
        await ctx.reply("✅ Removepoint dibatalkan.")
    else:
        await ctx.reply("❌ Tidak ada pending removepoint.")

@bot.command(name="ssrphelp")
async def cmd_help(ctx):
    embed = discord.Embed(title="🤖 SSRP Checker Bot", color=0x5865F2)
    embed.add_field(name="📋 Cara Submit", value=(
        "Kirim screenshot di:\n"
        "• `#ssrp-approval`\n"
        "• `#outrider-ssrp-approval`\n"
        "**1 foto = 1 poin** otomatis!"
    ), inline=False)
    embed.add_field(name="👤 Member", value=(
        "`!point` / `!point @user`\n"
        "`!profile` / `!profile @user`\n"
        "`!lb` — Leaderboard\n"
        "`!ssrphelp` — Bantuan"
    ), inline=False)
    embed.add_field(name="🔧 Admin", value=(
        "`!export excel/csv/all`\n"
        "`!inactive [hari]`\n"
        "`!resetweek` — Reset poin minggu semua\n"
        "`!resetuser @user` — Reset 1 member\n"
        "`!removeuser @user` — Hapus dari DB\n"
        "`!ekal` — **Simulasi reset Senin**"
    ), inline=False)
    embed.add_field(name="🚨 Darurat (saat bug)", value=(
        "`!addpoint @user [n]`\n"
        "`!removepoint @user [n]` + `!confirm` / `!cancel`"
    ), inline=False)
    embed.add_field(name="⚙️ Sistem", value=(
        f"• **1 foto = 1 poin**\n"
        f"• Cooldown: **{SUBMIT_COOLDOWN} detik** antar submit\n"
        f"• Reset otomatis setiap **Senin 00:00 WIB**\n"
        f"• Auto export Excel 3x sehari"
    ), inline=False)
    await ctx.reply(embed=embed)

# ─── Run ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    bot.run(DISCORD_TOKEN)
